#!/Users/azuev/projects/weekly_report/env/bin/python3
# -*- coding: utf-8 -*-
from datetime import datetime
import mail
from multiprocessing import Pool
import os

import openpyxl

import request_ga
import request_ym


def request_all(url_template):
    stat_g = request_ga.main_ga(url_template)
    stat_y = request_ym.main_ym(url_template)
    return stat_y, stat_g


time_begin = datetime.today()
# время запуска скрипта
time_begin_strf = time_begin.strftime('%d.%m.%Y %H:%M')
print(time_begin_strf)
# добавление времени запуска скрипта в словарь с результатами
dict_res_y = {'date': time_begin_strf}
dict_res_g = {'date': time_begin_strf}
list_url_with_date = ['date']

# чтение файла с url
with open('url.txt', 'r', encoding='utf-8') as r:
    content = r.read()
    list_url = content.split('\n')
    print(list_url)
    # создание объекта Pool на несколько процессов
    pool = Pool(processes=15)
    # запрос к api ga и ym с учетом мультипроцессинга
    res = pool.map(request_all, list_url)
    for url_idx, url in enumerate(list_url):
        # запись в словарь названия шаблона url и соответствующего значения по траифику из google за неделю
        dict_res_g[url] = res[url_idx][1]
        print(list_url[url_idx], 'g:', res[url_idx][1])
        # запись в словарь названия шаблона url и соответствующего значения по трафику из yandex за неделю
        dict_res_y[url] = res[url_idx][0]
        print(list_url[url_idx], 'y:', res[url_idx][0])
print(dict_res_y)
print(dict_res_g)
# формирования списка полей из шаблонов url и даты запуска
list_url_with_date = list_url_with_date + list_url
len_list = len(list_url_with_date)

# если файл отчета еще не существует, то создаем report.xlsx записываем результаты в 2 вкладки
if os.path.isfile('/Users/azuev/projects/weekly_report/report.xlsx') is False:
    workbook = openpyxl.Workbook()
    worksheet_y = workbook.active
    worksheet_y.title = 'Yandex'
    worksheet_g = workbook.create_sheet(title='Google')
    i = 0
    for idx_column in range(1, len_list + 1):
        worksheet_y.cell(row=1, column=idx_column).value = list_url_with_date[idx_column - 1]
        worksheet_y.cell(row=2, column=idx_column).value = dict_res_y[list_url_with_date[idx_column - 1]]
        worksheet_g.cell(row=1, column=idx_column).value = list_url_with_date[idx_column - 1]
        worksheet_g.cell(row=2, column=idx_column).value = dict_res_g[list_url_with_date[idx_column - 1]]
        workbook.save('report.xlsx')
# если файл отчета уже существует, то дозаписываем в него новые данные
else:
    workbook = openpyxl.load_workbook('report.xlsx')
    worksheet_y = workbook['Yandex']
    worksheet_g = workbook['Google']
    row_y = []
    row_g = []
    for url_idx in range(1, len_list + 1):
        row_y.append(dict_res_y[list_url_with_date[url_idx - 1]])
        row_g.append(dict_res_g[list_url_with_date[url_idx - 1]])
    worksheet_y.append(row_y)
    worksheet_g.append(row_g)
    workbook.save('report.xlsx')

# время окончания работы скрипта
time_end = datetime.today()
time_end_strf = time_end.strftime('%d.%m.%Y %H:%M')
print(time_end_strf)
delta = time_end - time_begin
# вывод длительности работы скрипта
print('время работы скрипта {0} минут {1} секунд'.format(delta.seconds // 60, delta.seconds % 60))

# mail.send_mail()
